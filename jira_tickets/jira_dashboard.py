import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


EXCEL_PATH = "jira_tickets.xlsx"
CSV_PATH = "jira_tickets.csv"
CURRENT_MONTH = datetime.now().strftime("%Y-%m")

# Carrega os dados da planilha para alimentar o dashboard
def load_source_data():

    if os.path.exists(EXCEL_PATH):
        try:
            wb = load_workbook(EXCEL_PATH, data_only=True)
        except PermissionError:
            raise SystemExit("Feche o arquivo jira_tickets.xlsx antes de executar o dashboard.")

        month_sheet = CURRENT_MONTH

        if month_sheet in wb.sheetnames:
            ws = wb[month_sheet]
            return pd.DataFrame(ws.values)

        data_sheets = [
            sheet for sheet in wb.sheetnames
            if sheet not in {"Resumo", "Status", "Assignee", "Sheet"}
        ]
        if data_sheets:
            ws = wb[data_sheets[-1]]
            return pd.DataFrame(ws.values)

    if os.path.exists(CSV_PATH):
        return pd.read_csv(CSV_PATH, encoding="utf-8-sig")

    raise FileNotFoundError("Nenhum arquivo de dados encontrado. Rode primeiro o script de exportação.")

# Padroniza o dataframe para o modelo de planilhas
def normalize_dataframe(raw_df):

    if raw_df is None or raw_df.empty:
        return pd.DataFrame(columns=["key", "summary", "status", "assignee", "created", "resolution"])

    df = raw_df.copy()
    if isinstance(df.columns, pd.Index) and len(df.columns) > 0:
        first_row = df.iloc[0].tolist()
        if any(str(value).lower() == "key" for value in first_row):
            df = df.iloc[1:].copy()
            df.columns = raw_df.iloc[0].tolist()

    expected = ["key", "summary", "status", "assignee", "created", "resolution"]
    for column in expected:
        if column not in df.columns:
            df[column] = ""

    df = df[expected].copy()
    return df.fillna("")


def build_summary_tables(df):
    """Cria os DataFrames de resumo por status e por responsável.

    O dashboard precisa mostrar rapidamente a carga de trabalho: total, abertos,
    fechados e distribuição por responsável para facilitar a leitura executiva.
    Também calcula percentuais para deixar a leitura mais rápida e direta.
    """
    total_tickets = len(df)
    open_tickets = df["status"].isin(["Open", "In Progress", "To Do"]).sum()
    resolved_tickets = df["status"].isin(["Resolved", "Closed", "Done"]).sum()

    status_summary = (
        df["status"].fillna("Sem status").value_counts().reset_index()
        .rename(columns={"index": "status", "status": "quantidade"})
        .sort_values("quantidade", ascending=False)
    )
    status_summary.columns = ["status", "quantidade"]
    if total_tickets:
        status_summary["percentual"] = (status_summary["quantidade"] / total_tickets * 100).round(1)
        status_summary["percentual_txt"] = status_summary["percentual"].map(lambda x: f"{x:.1f}%")
    else:
        status_summary["percentual"] = 0
        status_summary["percentual_txt"] = "0.0%"

    assignee_summary = (
        df["assignee"].fillna("Sem responsável").value_counts().reset_index()
        .rename(columns={"index": "assignee", "assignee": "quantidade"})
        .sort_values("quantidade", ascending=False)
    )
    assignee_summary.columns = ["assignee", "quantidade"]
    if total_tickets:
        assignee_summary["percentual"] = (assignee_summary["quantidade"] / total_tickets * 100).round(1)
        assignee_summary["percentual_txt"] = assignee_summary["percentual"].map(lambda x: f"{x:.1f}%")
    else:
        assignee_summary["percentual"] = 0
        assignee_summary["percentual_txt"] = "0.0%"

    overview = pd.DataFrame(
        {
            "metric": [
                "Total de tickets",
                "Tickets abertos",
                "Tickets resolvidos/fechados",
                "% abertos",
                "% resolvidos/fechados",
            ],
            "valor": [
                total_tickets,
                open_tickets,
                resolved_tickets,
                f"{(open_tickets / total_tickets * 100) if total_tickets else 0:.1f}%",
                f"{(resolved_tickets / total_tickets * 100) if total_tickets else 0:.1f}%",
            ],
        }
    )

    return overview, status_summary, assignee_summary


def apply_dashboard_style(ws):
    """Aplica a formatação visual padrão para as abas do dashboard.

    Essa rotina garante que o Excel final tenha aparência limpa e organizada,
    mantendo texto centralizado, colunas ajustadas e cabeçalhos consistentes.
    """
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 15), 35)


def write_dashboard_sheets(df, workbook_path):
    """Escreve as abas de resumo no mesmo arquivo Excel.

    A ideia é manter o workbook principal e adicionar abas separadas para facilitar
    a leitura de gestão, sem criar um arquivo adicional e sem quebrar o histórico.
    """
    if not os.path.exists(workbook_path):
        wb = Workbook()
    else:
        try:
            wb = load_workbook(workbook_path)
        except PermissionError:
            raise SystemExit("Feche o arquivo jira_tickets.xlsx antes de executar o dashboard.")

    for sheet_name in ["Resumo", "Status", "Assignee", "Gráficos", "Gráficos1"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    overview_sheet = wb.create_sheet("Resumo")
    status_sheet = wb.create_sheet("Status")
    assignee_sheet = wb.create_sheet("Assignee")

    overview, status_summary, assignee_summary = build_summary_tables(df)

    overview_sheet.append(["Métrica", "Valor"])
    for row in overview.itertuples(index=False):
        overview_sheet.append(list(row))

    status_sheet.append(["Status", "Quantidade", "% do total"])
    for row in status_summary.itertuples(index=False):
        status_sheet.append([row.status, row.quantidade, row.percentual_txt])

    assignee_sheet.append(["Assignee", "Quantidade", "% do total"])
    for row in assignee_summary.itertuples(index=False):
        assignee_sheet.append([row.assignee, row.quantidade, row.percentual_txt])

    for ws in [overview_sheet, status_sheet, assignee_sheet]:
        apply_dashboard_style(ws)

    chart_sheet = wb.create_sheet("Gráficos")
    chart_sheet["A1"] = ""
    chart_sheet["A1"].font = Font(bold=True, size=14, color="1F4E78")

    if not status_summary.empty:
        chart_sheet["A3"] = "Status"
        chart_sheet["A3"].font = Font(bold=True)
        chart_sheet["B3"] = "Quantidade"
        chart_sheet["C3"] = "%"
        chart_sheet["B3"].font = Font(bold=True)
        chart_sheet["C3"].font = Font(bold=True)

        for row_idx, row in enumerate(status_summary.itertuples(index=False), start=4):
            chart_sheet[f"A{row_idx}"] = row.status
            chart_sheet[f"B{row_idx}"] = row.quantidade
            chart_sheet[f"C{row_idx}"] = row.percentual_txt

        status_data = Reference(chart_sheet, min_col=2, min_row=3, max_row=3 + len(status_summary))
        status_cats = Reference(chart_sheet, min_col=1, min_row=4, max_row=3 + len(status_summary))

        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.style = 10
        bar_chart.title = "Tickets por status"
        bar_chart.y_axis.title = "Quantidade"
        bar_chart.x_axis.title = "Status"
        bar_chart.height = 7
        bar_chart.width = 12
        bar_chart.add_data(status_data, titles_from_data=False)
        bar_chart.set_categories(status_cats)
        chart_sheet.add_chart(bar_chart, "E4")

    if not assignee_summary.empty:
        chart_sheet["H3"] = "Responsável"
        chart_sheet["I3"] = "Quantidade"
        chart_sheet["J3"] = "%"
        chart_sheet["H3"].font = Font(bold=True)
        chart_sheet["I3"].font = Font(bold=True)
        chart_sheet["J3"].font = Font(bold=True)

        rows_to_write = assignee_summary.head(8).itertuples(index=False)
        for row_idx, row in enumerate(rows_to_write, start=4):
            chart_sheet[f"H{row_idx}"] = row.assignee
            chart_sheet[f"I{row_idx}"] = row.quantidade
            chart_sheet[f"J{row_idx}"] = row.percentual_txt

        pie_data = Reference(chart_sheet, min_col=9, min_row=3, max_row=3 + min(len(assignee_summary), 8))
        pie_cats = Reference(chart_sheet, min_col=8, min_row=4, max_row=3 + min(len(assignee_summary), 8))

        pie_chart = PieChart()
        pie_chart.title = "Distribuição por responsável"
        pie_chart.height = 7
        pie_chart.width = 12
        pie_chart.add_data(pie_data, titles_from_data=False)
        pie_chart.set_categories(pie_cats)
        chart_sheet.add_chart(pie_chart, "P4")

    apply_dashboard_style(chart_sheet)
    wb.save(workbook_path)


def main():
    """Ponto de entrada do dashboard.

    O fluxo principal lê a planilha do mês atual, normaliza os dados e grava as
    abas de resumo no mesmo workbook para uso executivo e operacional.
    """
    raw_df = load_source_data()
    df = normalize_dataframe(raw_df)
    write_dashboard_sheets(df, EXCEL_PATH)

    print("Dashboard atualizado com sucesso.")
    print(f"Arquivo processado: {EXCEL_PATH}")


if __name__ == "__main__":
    main()
