import os
from datetime import datetime, timedelta

import pandas as pd
import requests
import urllib3
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


ASSIGNEE_IDS = [
    "712020:16b30358-f02a-4f08-8783-910e6d4d55aa",
    "712020:87f53f74-c669-4e5e-997c-66f6325cb733",
    "712020:eb0255e3-2175-4f95-b860-535b9ac24a8d",
    "712020:0d271e63-4899-4b42-bb0a-37a3bd2c5507",
    "712020:636e3fd7-2309-41a1-a3c0-a3e1eaec36a0",
    "712020:06f03753-08f6-4477-b2f3-2d39d687e8fa",
]
FIELD_NAMES = ["summary", "status", "assignee", "created", "resolution"]
COLUMNS = ["key", "summary", "status", "assignee", "created", "resolution"]

# Carrega as credenciais do Jira a partir do .env    
def load_jira_credentials():
    
    jira_url = os.getenv("JIRA_URL", "https://seu-dominio.atlassian.net").rstrip("/")
    email = os.getenv("JIRA_EMAIL", "")
    token = os.getenv("JIRA_TOKEN", "")

    if not email or not token:
        raise SystemExit("Configure JIRA_EMAIL e JIRA_TOKEN antes de executar o script.")

    return jira_url, email, token

# Divide o período em janelas de datas. a busca é feita por intervalos de tempo
# para manter a paginação funcional sem quebrar a api    
def build_date_windows(start_date, end_date, chunk_days=7):
    
    current = start_date
    while current <= end_date:
        window_end = min(current + timedelta(days=chunk_days), end_date)
        yield current, window_end
        current = window_end + timedelta(days=1)

# Montagem do filtro de tickets JQL para cada janela de data.
def build_jql(window_start, window_end, assignee_ids=None):
    
    assignee_ids = assignee_ids or ASSIGNEE_IDS
    assignee_filter = ", ".join(f"{uid}" for uid in assignee_ids)
    return f'''
        created >= "{window_start.strftime('%Y-%m-%d')}"
        AND project = HELP
        AND assignee IN ({assignee_filter})
        AND created <= "{window_end.strftime('%Y-%m-%d')}"
        ORDER BY created DESC
    '''

# Executa a chamada HTTP para a API do Jira.
def request_search(jira_url, auth, jql, fields, max_results=100, start_at=None):
    

    payload = {
        "jql": jql,
        "fields": fields,
        "maxResults": max_results,
    }
    if start_at is not None:
        payload["startAt"] = start_at

    response = requests.post(
        f"{jira_url}/rest/api/3/search/jql",
        auth=auth,
        json=payload,
        timeout=40,
        verify=False,
    )

    if response.status_code != 200:
        text = response.text[:500]
        if start_at is not None and "startAt" in text.lower():
            return None
        print("Erro:", response.status_code)
        print(text)
        raise SystemExit("Erro ao buscar tickets no Jira.")

    return response.json()


# Transforma dados brutos do Jira em dicionário pronto para exportar como tabela 
def normalize_issue(issue, assignee_ids):
   
    fields = issue.get("fields", {})
    resolution = fields.get("resolution") or {}
    assignee = fields.get("assignee") or {}
    status = fields.get("status") or {}

    assigned_id = assignee.get("accountId") or assignee.get("key") or assignee.get("name") or ""
    if assigned_id and assigned_id not in assignee_ids:
        return None

    return {
        "key": issue.get("key", ""),
        "summary": fields.get("summary", ""),
        "status": status.get("name", ""),
        "assignee": assignee.get("displayName", ""),
        "created": fields.get("created", ""),
        "resolution": resolution.get("name", ""),
    }

# Busca os tickets dentro do filtro de data, semana a semana até pegar todos.
def collect_rows(jira_url, auth, assignee_ids, month_start, month_end, target_tickets, fields):

    rows = []
    for window_start, window_end in build_date_windows(month_start, month_end, chunk_days=7):
        jql = build_jql(window_start, window_end, assignee_ids)

        data = request_search(jira_url, auth, jql, fields, max_results=100)
        if data is None:
            data = request_search(jira_url, auth, jql, fields, max_results=100, start_at=None)

        issues = data.get("issues", [])
        if not issues:
            continue

        print(f"Janela {window_start.strftime('%Y-%m-%d')} a {window_end.strftime('%Y-%m-%d')}: {len(issues)} issue(s) retornadas.")

        for issue in issues:
            row = normalize_issue(issue, assignee_ids)
            if row is not None:
                rows.append(row)

            if len(rows) >= target_tickets:
                return rows

        if len(rows) >= target_tickets:
            break

    return rows
# Estilo padronizado para as planilhas :)
def apply_sheet_style(ws):

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in column_cells
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 18), 40)

    if ws.max_row > 1:
        status_colors = {
            "Open": "FFF2CC",
            "In Progress": "D9EAF7",
            "Resolved": "D9EAD3",
            "Closed": "E2F0D9",
            "To Do": "FCE5CD",
            "Done": "D9EAD3",
        }
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            status_value = str(row[2].value or "").strip()
            if status_value in status_colors:
                row[2].fill = PatternFill(fill_type="solid", fgColor=status_colors[status_value])


# Salva os dados em um arquivo Excel organizado por abas separadas por mês 
def save_excel_with_month_sheet(df, excel_path, month_sheet):

    if os.path.exists(excel_path):
        try:
            wb = load_workbook(excel_path)
        except PermissionError:
            raise SystemExit("Feche o arquivo jira_tickets.xlsx antes de executar o script.")
    else:
        wb = Workbook()

    if month_sheet in wb.sheetnames:
        ws = wb[month_sheet]
        existing_keys = {
            ws.cell(row=row_idx, column=1).value
            for row_idx in range(2, ws.max_row + 1)
            if ws.cell(row=row_idx, column=1).value is not None
        }

        for row in df.itertuples(index=False):
            if row[0] not in existing_keys:
                ws.append(row)
                existing_keys.add(row[0])
    else:
        ws = wb.active if wb.sheetnames else wb.create_sheet(title=month_sheet)
        if wb.sheetnames and wb.sheetnames[0] == "Sheet":
            ws.title = month_sheet
        else:
            wb.create_sheet(title=month_sheet)
            ws = wb[month_sheet]

        ws.append(COLUMNS)
        for row in df.itertuples(index=False):
            ws.append(row)

    if not df.empty and ws.max_row == 1:
        ws.append(COLUMNS)

    apply_sheet_style(ws)
    wb.save(excel_path)

# Autentica, busca, transforma em DataFrame, salva em .xlsx e .csv e retorna dados no terminal
def main():

    jira_url, email, token = load_jira_credentials()
    auth = (email, token)

    today = datetime.today()
    month_start = today.replace(day=1)
    month_end = today

    target_tickets = int(os.getenv("JIRA_TARGET_TICKETS", "500"))
    rows = collect_rows(
        jira_url=jira_url,
        auth=auth,
        assignee_ids=ASSIGNEE_IDS,
        month_start=month_start,
        month_end=month_end,
        target_tickets=target_tickets,
        fields=FIELD_NAMES,
    )

    df = pd.DataFrame(rows, columns=COLUMNS)
    excel_path = "jira_tickets.xlsx"
    csv_path = "jira_tickets.csv"
    month_sheet = today.strftime("%Y-%m")

    save_excel_with_month_sheet(df, excel_path, month_sheet)
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    print(f"Total de issues: {len(rows)}")
    print(f"Mês atual: {month_sheet}")
    print(f"Arquivo Excel salvo: {excel_path}")
    print(f"Arquivo CSV salvo: {csv_path}")
    print(df.head().to_string(index=False))


if __name__ == "__main__":
    main()

    